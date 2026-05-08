import datetime
import os 
import csv
import json
from decimal import Decimal
from urllib.parse import quote
from decimal import InvalidOperation
from django.utils import timezone
from rest_framework import viewsets, status, generics
from rest_framework.decorators import action
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.exceptions import PermissionDenied, ValidationError
from django.db import transaction
from django.db import models
from django.db.models import Count, Q
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from django.template.loader import render_to_string
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.parsers import MultiPartParser, FormParser
from django.db.models import Case, When, Value, IntegerField
from users import permissions
from .serializers import DoctorSerializer
from .models import Doctor
from .models import (
    Patient,
    Admission,
    MedicalHistory,
    Discharge,
    Service,
    Billing,
    ServiceMaster,
    DischargeSummary,
    Task,
    LabReport,
    HODReview,
    DepartmentLogEntry,
    HospitalSettings,
)
from .serializers import (
    PatientSerializer,
    MedicalHistorySerializer,
    DischargeSerializer,
    ServiceSerializer,
    BillingSerializer,
    ServiceMasterSerializer,
    DischargeSummarySerializer,
    TaskSerializer,
    LabReportSerializer,
    HODReviewSerializer,
    DepartmentLogEntrySerializer, BulkTaskAssignSerializer, HospitalSettingsSerializer
)
from .report_templates import build_suggested_reports_for_admission
import qrcode
import base64
import io
import openpyxl
from django.conf import settings
from django.template.loader import render_to_string
from django.http import HttpResponse
from xhtml2pdf import pisa
import copy
from .templates import DISCHARGE_TEMPLATES
import io
from xhtml2pdf import pisa
from users.models import CustomUser
from .models import MedicineMaster, PharmacyRecord, ReportMaster
from .serializers import MedicineMasterSerializer, PharmacyRecordSerializer, ReportMasterSerializer

DEPARTMENT_ROLE_MAP = {
    'HOD': 'hod',
    'Billing': 'billing',
    'Uploading': 'uploading',
    'Query': 'query',
    'OPD': 'opd',
    'Intimation': 'intimation',
    'Receptionist': 'receptionist',
    'Nursing': 'nursing',
    'Doctor': 'doctor',
    'Notes': 'notes',
    'Quality Analysis': 'quality_analyst',  
}

DEPARTMENT_LOG_FIELDS = {
    'opd': ['uploadDate', 'createdAt', 'opdDate'],
    'intimation': ['uploadDate', 'createdAt', 'doa'],
    'query': ['queryRepDate', 'createdAt', 'raiseDate'],
    'uploading': ['uploadDate', 'createdAt', 'doa'],
}


def get_branch_settings_queryset():
    return HospitalSettings.objects.all().order_by('branch_name', 'branch')


def get_valid_branch_codes():
    return set(get_branch_settings_queryset().values_list('branch', flat=True))


def resolve_branch_code_from_loc(loc_id=None, explicit_branch=None):
    if explicit_branch:
        branch = str(explicit_branch).strip().upper()
        if HospitalSettings.objects.filter(branch=branch).exists():
            return branch

    if loc_id:
        slug = str(loc_id).strip().lower()
        branch_obj = HospitalSettings.objects.filter(slug=slug).first()
        if branch_obj:
            return branch_obj.branch

    default_branch = HospitalSettings.objects.order_by('id').first()
    return default_branch.branch if default_branch else 'LNM'
TASK_MANAGER_ROLES = {'superadmin', 'office_admin', 'admin', 'hod'}
TASK_ASSIGNABLE_ROLES = {
    'receptionist', 'billing', 'hod', 'opd', 'intimation', 'query', 'uploading',
    'nursing', 'notes', 'medical_officer', 'quality_analyst',
}

def normalize_task_status(raw_status, due_date=None):
    status_map = {
        'pending': 'Pending',
        'in-progress': 'In Progress',
        'completed': 'Completed',
        'on-hold': 'On Hold',
        'overdue': 'Overdue',
    }
    safe_status = status_map.get(str(raw_status or '').strip().lower(), 'Pending')
    if safe_status != 'Completed' and due_date and due_date < timezone.now():
        return 'Overdue'
    return safe_status

def serialize_task_for_hod(task):
    patient = task.patient
    status_value = task.status
    if status_value != 'Completed' and task.due_date and task.due_date < timezone.now():
        status_value = 'Overdue'

    status_map = {
        'Pending': 'pending',
        'In Progress': 'in-progress',
        'Completed': 'completed',
        'On Hold': 'pending',
        'Overdue': 'overdue',
    }
    priority_map = {
        'Low': 'low',
        'Medium': 'medium',
        'High': 'high',
        'Urgent': 'high',
    }

    employee_name = task.assigned_to.get_full_name().strip() or task.assigned_to.username

    patient_type = 'TPA'
    if patient:
        if (patient.cashlessType or '').lower().find('card') >= 0:
            patient_type = 'Card'
        elif (patient.payMode or '').lower().find('cash') >= 0:
            patient_type = 'Cash'

    return {
        'id': task.id,
        'employeeId': task.assigned_to_id,
        'employeeName': employee_name,
        'taskType': task.title,
        'patientId': patient.uhid if patient else '',
        'patientType': patient_type,
        'priority': priority_map.get(task.priority, 'medium'),
        'dueDate': task.due_date.date().isoformat() if task.due_date else '',
        'status': status_map.get(status_value, 'pending'),
        'notes': task.description or '',
        'department': task.department,
    }

def get_task_queryset_for_user(user):
    queryset = (
        Task.objects
        .select_related('assigned_to', 'assigned_by', 'patient')
        .prefetch_related(
            'patient__admissions',
            'patient__admissions__medicalHistory',
            'patient__admissions__discharge',
            'patient__admissions__services',
            'patient__admissions__bills',
            'patient__admissions__lab_reports',
            'patient__admissions__pharmacy_records',
        )
        .order_by('-created_at')
    )
    if user.role in ['superadmin', 'office_admin']:
        return queryset
    if user.role == 'admin':
        return queryset.filter(
            models.Q(assigned_to__branch=user.branch) |
            models.Q(assigned_by=user)
        )
    if user.role == 'hod':
        return queryset.filter(models.Q(assigned_to=user) | models.Q(assigned_by=user))
    return queryset.filter(assigned_to=user)

def validate_generic_task_assignment(actor, assigned_to, patient=None, department=None):
    valid_branch_codes = get_valid_branch_codes()
    if actor.role not in TASK_MANAGER_ROLES:
        raise PermissionDenied("You are not allowed to manage tasks from this dashboard.")

    if actor.role == 'superadmin':
        allowed_roles = TASK_ASSIGNABLE_ROLES | {'admin', 'office_admin'}
    else:
        allowed_roles = TASK_ASSIGNABLE_ROLES

    if assigned_to.role not in allowed_roles:
        raise PermissionDenied(
            f"{actor.get_role_display()} cannot assign tasks to {assigned_to.get_role_display()} accounts."
        )
    
    if (
        patient and
        actor.role not in {'office_admin', 'superadmin', 'hod'} and 
        assigned_to.branch in valid_branch_codes and
        patient.branch_location != assigned_to.branch
    ):
        raise ValidationError({'patient': 'Selected patient must belong to the same branch as the assigned employee.'})

    expected_role = get_department_role(department)
    if expected_role and assigned_to.role != expected_role:
        raise ValidationError({
            'assigned_to': f"Department '{department}' tasks must be assigned to a '{expected_role}' user."
        })

    if expected_role == 'billing' and patient is None:
        raise ValidationError({'patient': 'Billing tasks must be linked to a patient.'})

    if actor.role == 'admin' and assigned_to.branch != actor.branch:
        raise PermissionDenied("Branch Admin can assign tasks only inside their own branch.")

    if (
        patient and
        actor.role not in {'office_admin', 'superadmin'} and
        assigned_to.branch in valid_branch_codes and
        patient.branch_location != assigned_to.branch
    ):
        raise ValidationError({'patient': 'Selected patient must belong to the same branch as the assigned employee.'})

def get_department_role(department):
    return DEPARTMENT_ROLE_MAP.get(str(department or '').strip(), '')

def get_allowed_hod_departments(user):
    if user.role == 'superadmin':
        return list(DEPARTMENT_ROLE_MAP.keys())
    if user.role == 'office_admin':
        return list(DEPARTMENT_ROLE_MAP.keys())
    if user.role == 'hod':
        return list(DEPARTMENT_ROLE_MAP.keys())
    return []

def ensure_hod_access(request):
    role = getattr(request.user, 'role', '')
    if not request.user.is_authenticated or role not in ['hod', 'office_admin', 'superadmin']:
        return Response({'error': 'Unauthorized access.'}, status=status.HTTP_403_FORBIDDEN)
    return None

def coerce_record_date(department, payload):
    for key in DEPARTMENT_LOG_FIELDS.get(department, []):
        value = payload.get(key)
        if not value:
            continue
        if isinstance(value, str):
            safe_value = value[:10]
            try:
                return datetime.date.fromisoformat(safe_value)
            except ValueError:
                continue
    return timezone.localdate()

def get_or_create_current_billing(admission):
    billing = admission.bills.order_by('-id').first()
    if billing:
        return billing, False
    return Billing.objects.create(admission=admission), True

def normalize_service_pricing(service_data, patient=None):
    raw_pricing = str(
        service_data.get('pricing_type')
        or service_data.get('pricingApplied')
        or service_data.get('pricing_applied')
        or ''
    ).strip().upper()
    if raw_pricing in {'CASH', 'CASHLESS'}:
        return raw_pricing
    pay_mode = str(getattr(patient, 'payMode', '') or '').lower()
    return 'CASHLESS' if 'cashless' in pay_mode else 'CASH'

def resolve_service_defaults(service_data, patient=None):
    svc_name = (service_data.get('svcName') or service_data.get('title') or service_data.get('name') or '').strip()
    if not svc_name:
        raise ValueError('Service name (svcName) is required.')

    pricing_applied = normalize_service_pricing(service_data, patient)
    svc_date = service_data.get('svcDate') or service_data.get('date') or None
    
    try:
        svc_qty = int(service_data.get('svcQty') or service_data.get('qty') or 1)
    except (TypeError, ValueError):
        svc_qty = 1
    svc_qty = max(1, svc_qty)

    master_service = ServiceMaster.objects.filter(
        description__iexact=svc_name,
        pricing_type=pricing_applied,
    ).first()

    if master_service:
        svc_rate = master_service.rate
        svc_cat = master_service.category
        svc_code = master_service.code  # 🌟 NEW: Grab the code from the Excel Master!
    else:
        raw_rate = service_data.get('svcRate') or service_data.get('rate') or 0
        raw_cat = service_data.get('svcCat') or service_data.get('type') or 'GENERAL SERVICES'
        svc_code = service_data.get('svcCode') or service_data.get('code') or '' # 🌟 NEW: Fallback
        try:
            svc_rate = Decimal(str(raw_rate))
        except (InvalidOperation, ValueError, TypeError):
            svc_rate = Decimal('0')
        svc_cat = raw_cat

    return {
        'svcName': svc_name,
        'svcCode': svc_code, # 🌟 NEW: Return it so the view can save it!
        'pricing_applied': pricing_applied,
        'svcCat': svc_cat,
        'svcQty': svc_qty,
        'svcRate': svc_rate,
        'svcTot': svc_rate * svc_qty,
        'svcDate': svc_date,
    }

class PatientViewSet(viewsets.ModelViewSet):
    queryset = Patient.objects.all().order_by('-created_at')
    serializer_class = PatientSerializer
    lookup_field = 'uhid'
    lookup_value_regex = '[^/]+'

    def get_queryset(self):
        user = self.request.user

        if not getattr(user, 'is_authenticated', False):
            return Patient.objects.none()

        queryset = Patient.objects.all()

        # 🌟 THE TASK FIX: Only hides patients if the frontend explicitly asks (for the Assignment Modal)
        exclude_dept = self.request.query_params.get('exclude_active_tasks_for_dept')
        if exclude_dept:
            queryset = queryset.exclude(
                assigned_tasks__department__iexact=exclude_dept,
                assigned_tasks__status__in=['Pending', 'In Progress']
            )

        # 1. 🌍 Super Admin & Office Admin: See EVERYTHING across ALL branches
        if user.role in ['superadmin', 'office_admin']:
            return queryset.order_by('-created_at')

        # 2. 🏥 Branch Admin & Receptionist: See ALL patients for THEIR branch
        elif user.role in ['admin', 'receptionist']:
            return queryset.filter(branch_location=user.branch).order_by('-created_at')

        # 3. 👔 HOD: Sees ALL CASHLESS patients (all hospitals) + Tasks assigned to/by them
        elif user.role == 'hod':
            from django.db import models
            return queryset.filter(
                models.Q(assigned_tasks__assigned_to=user) | 
                models.Q(assigned_tasks__assigned_by=user) |
                models.Q(payMode__icontains='cashless') |
                models.Q(admissions__bills__bill_type='CASHLESS')
            ).distinct().order_by('-created_at')

        # 4. 👩‍⚕️ Staff (Created by Office Admin/HOD): See ONLY patients explicitly assigned to them
        elif user.role in ['billing', 'opd', 'intimation', 'query', 'uploading', 'nursing', 'notes', 'medical_officer', 'quality_analyst']:
            return queryset.filter(assigned_tasks__assigned_to=user).distinct().order_by('-created_at')

        return queryset.none()
    
    def create(self, request, *args, **kwargs):
        data = request.data.copy()
        admission_type = data.pop('admissionType', None) or request.data.get('admissionType') or 'IPD'

         # If no branch sent, use the logged-in user's branch (not DB first branch)
        if not data.get('branch_location') and not data.get('locId'):
            if getattr(request.user, 'branch', None) not in [None, 'ALL']:
                data['branch_location'] = request.user.branch
        
        if 'locId' in data or 'branch_location' in data:
            data['branch_location'] = resolve_branch_code_from_loc(data.get('locId'), data.get('branch_location'))
            
        serializer = self.get_serializer(data=data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        patient = serializer.instance 

        try:
            Admission.objects.create(patient=patient, admNo=1, admissionType=admission_type)
            response_serializer = self.get_serializer(patient)
            headers = self.get_success_headers(response_serializer.data)
            return Response(response_serializer.data, status=status.HTTP_201_CREATED, headers=headers)
        except Exception as e:
            print("🚨 AUTO-ADMISSION FAILED:", str(e))
            return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'], url_path='new_admission')
    def new_admission(self, request, uhid=None): 
        try:
            # 🌟 THE FIX: Fetch the patient manually using the UHID from the URL
            # This guarantees it won't fail trying to find a default 'pk'
            patient = get_object_or_404(Patient, uhid=uhid)
            
            admission_type = request.data.get('admissionType') or 'IPD'
            
            last_adm = Admission.objects.filter(patient=patient).order_by('id').last()
            new_adm_no = (last_adm.admNo + 1) if last_adm else 1
            
            admission = Admission.objects.create(
                patient=patient, 
                admNo=new_adm_no,
                admissionType=admission_type,
            )
            
            # Return the fully updated patient profile to the frontend
            serializer = self.get_serializer(patient)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            print("🚨 ADMISSION CREATION FAILED:", str(e))
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
               

    @action(detail=True, methods=['patch'])
    def update_medical(self, request, uhid=None):
        patient = self.get_object()
        adm_no = request.data.get('admNo') or 1
        medical_data = request.data.get('medicalData', {})
        
        try:
            admission_obj, _ = Admission.objects.get_or_create(patient=patient, admNo=adm_no)
            med_hist, _ = MedicalHistory.objects.get_or_create(admission=admission_obj)
                
            for key, value in medical_data.items():
                if key in ['id', 'admission']:
                    continue
                setattr(med_hist, key, value)
                
            med_hist.save()
            return Response({'status': 'Medical history updated successfully'})
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['patch'])
    def discharge(self, request, uhid=None):
        patient = self.get_object()
        adm_no = request.data.get('admNo') or 1
        discharge_data = request.data.get('dischargeData', {})
        
        try:
            admission_obj, _ = Admission.objects.get_or_create(patient=patient, admNo=adm_no)
            discharge_obj, _ = Discharge.objects.get_or_create(admission=admission_obj)
                
            for key, value in discharge_data.items():
                if key in ['id', 'admission']: 
                    continue
                if key in ['dod', 'expectedDod', 'actualDod'] and value == "":
                    value = None
                setattr(discharge_obj, key, value)
                
            discharge_obj.save()
            return Response({'status': 'Discharge updated successfully'})
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        
    @action(detail=True, methods=['patch'])
    def update_billing(self, request, uhid=None):
        patient = self.get_object()
        adm_no = request.data.get('admNo') or 1
        billing_data = request.data.get('billingData', {})
        
        try:
            admission_obj, _ = Admission.objects.get_or_create(patient=patient, admNo=adm_no)
            billing_obj, _ = get_or_create_current_billing(admission_obj)
                
            for key, value in billing_data.items():
                # 🌟 PROTECT STATUS: Ignore printStatus so staff saves don't overwrite Admin approvals!
                if key in ['id', 'admission', 'printStatus', 'printRequestedAt']: 
                    continue
                    
                if key in ['discount', 'advance', 'paidNow']:
                    if value in ["", None]:
                        value = Decimal('0')
                    else:
                        try:
                            value = Decimal(str(value))
                        except (InvalidOperation, ValueError, TypeError):
                            value = Decimal('0')
                    
                setattr(billing_obj, key, value)

            pay_mode = str(billing_data.get('paymentMode') or getattr(billing_obj, 'paymentMode', '') or '')
            insurance_type = str(billing_data.get('insuranceType') or getattr(billing_obj, 'insuranceType', '') or '')
            cashless_like = {'tpa', 'echs', 'eci', 'fci', 'ayushman bharat', 'northern railways', 'insurance'}
            billing_obj.bill_type = 'CASHLESS' if (
                'cashless' in pay_mode.lower() or insurance_type.strip().lower() in cashless_like
            ) else 'CASH'
                
            billing_obj.save()
            return Response({'status': 'Billing updated successfully'})
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['post'])
    def add_service(self, request, uhid=None):
        patient = self.get_object()
        adm_no = request.data.get('admNo') or 1
        service_data = request.data.get('serviceData')
        
        try:
            admission_obj, _ = Admission.objects.get_or_create(patient=patient, admNo=adm_no)
            defaults = resolve_service_defaults(service_data or {}, patient)
            service, created = Service.objects.update_or_create(
                admission=admission_obj,
                svcName=defaults['svcName'],
                pricing_applied=defaults['pricing_applied'],
                defaults={
                    'svcCat': defaults['svcCat'],
                    'svcQty': defaults['svcQty'],
                    'svcRate': defaults['svcRate'],
                    'svcTot': defaults['svcTot'],
                    'svcDate': defaults['svcDate'],
                    'svcCode': defaults['svcCode'],
                }
            )
            
            return Response({
                'status': 'Service added successfully with automated pricing.', 
                'data': ServiceSerializer(service).data
            })
            
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        
    @action(detail=True, methods=['patch'])
    def set_expected_dod(self, request, uhid=None):
        patient = self.get_object()
        adm_no = request.data.get('admNo')
        expected_date = request.data.get('expectedDod')
        
        if expected_date == "":
            expected_date = None
        elif expected_date and len(expected_date) > 10:
            expected_date = expected_date[:10]

        try:
            admission = patient.admissions.get(admNo=adm_no)
            if not hasattr(admission, 'discharge'):
                Discharge.objects.create(admission=admission)
            
            admission.discharge.expectedDod = expected_date
            admission.discharge.save()
            return Response({'status': 'Expected DOD updated successfully'})
        except Exception as e:
            print("🚨 DOD ERROR:", str(e))
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'])
    def request_print(self, request, uhid=None):
        patient = self.get_object()
        adm_no = request.data.get('admNo')
        
        try:
            admission = patient.admissions.get(admNo=adm_no)
            billing_obj, _ = get_or_create_current_billing(admission)
            
            # 🌟 SMART CHECK: Prevent resetting an already approved bill back to PENDING!
            if billing_obj.printStatus == 'APPROVED':
                return Response({'status': 'Already approved'})

            billing_obj.printStatus = 'PENDING'
            billing_obj.printRequestedAt = timezone.now()
            billing_obj.save()
            return Response({'status': 'Print request sent to Super Admin'})
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'])
    def resolve_print(self, request, uhid=None):
        patient = self.get_object()
        adm_no = request.data.get('admNo')
        
        action = request.data.get('action') or request.data.get('status') or request.data.get('backendAction') or 'APPROVED'
        
        try:
            admission = patient.admissions.get(admNo=adm_no)
            billing_obj, _ = get_or_create_current_billing(admission)
            billing_obj.printStatus = action
            billing_obj.save()
                
            return Response({'status': f'Print request {action}'})
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'])
    def pending_prints(self, request):
        pending_patients = Patient.objects.filter(admissions__bills__printStatus='PENDING').distinct()
        
        serializer = self.get_serializer(pending_patients, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'], url_path='cashless-records')
    def cashless_records(self, request):
        # 1. Strict Security Check: Only Office Admins can hit this endpoint
        if getattr(request.user, 'role', '') != 'office_admin':
            return Response(
                {"error": "Unauthorized access. Only Office Admins can view the corporate dashboard."}, 
                status=status.HTTP_403_FORBIDDEN
            )
        
        # 2. Database Query: Find patients linked to an admission that has a cashless bill
        # The .distinct() ensures we don't get duplicate patients if they have multiple cashless visits
        from .models import Patient
        cashless_patients = Patient.objects.filter(admissions__bills__bill_type='CASHLESS').distinct()
        
        # 3. Serialize and Return
        # Because the user is 'office_admin', our updated serializer will naturally 
        # expose all the prices and totals we hid from the hospital staff!
        serializer = self.get_serializer(cashless_patients, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

class ServiceBulkSaveAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, uhid, adm_no):
        patient = get_object_or_404(Patient, uhid=uhid)
        admission_obj, _ = Admission.objects.get_or_create(patient=patient, admNo=adm_no)
        services = request.data.get('services') or []

        if not isinstance(services, list):
            return Response({'error': 'services must be a list.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            serialized = []
            with transaction.atomic():
                admission_obj.services.all().delete()
                created_services = []
                for service_data in services:
                    defaults = resolve_service_defaults(service_data or {}, patient)
                    created_services.append(Service(
                        admission=admission_obj,
                        svcName=defaults['svcName'],
                        svcCode=defaults['svcCode'],  # 🌟 NEW: Saving the Code!
                        pricing_applied=defaults['pricing_applied'],
                        svcCat=defaults['svcCat'],
                        svcQty=defaults['svcQty'],
                        svcRate=defaults['svcRate'],
                        svcTot=defaults['svcTot'],
                        svcDate=defaults['svcDate'],
                    ))
                if created_services:
                    Service.objects.bulk_create(created_services)
                serialized = ServiceSerializer(admission_obj.services.order_by('svcDate', 'id'), many=True).data
            return Response({'saved': len(serialized), 'services': serialized}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
    
class ServiceMasterViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = ServiceMasterSerializer
    pagination_class = None

    def get_queryset(self):
        qs = ServiceMaster.objects.all()
        pricing = self.request.query_params.get('pricing_type')
        if pricing:
            qs = qs.filter(pricing_type=pricing.upper())
        return qs


class HospitalSettingsViewSet(viewsets.ModelViewSet):
    queryset = get_branch_settings_queryset()
    serializer_class = HospitalSettingsSerializer
    pagination_class = None

    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            return [AllowAny()]
        return [IsAuthenticated()]

    def create(self, request, *args, **kwargs):
        if getattr(request.user, 'role', '') != 'superadmin':
            raise PermissionDenied("Only Super Admin can create hospital branches.")
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        if getattr(request.user, 'role', '') != 'superadmin':
            raise PermissionDenied("Only Super Admin can update hospital branches.")
        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        if getattr(request.user, 'role', '') != 'superadmin':
            raise PermissionDenied("Only Super Admin can delete hospital branches.")
        instance = self.get_object()
        if Patient.objects.filter(branch_location=instance.branch).exists():
            raise ValidationError({'branch': 'This branch already has patient records and cannot be deleted.'})
        if CustomUser.objects.filter(branch=instance.branch).exists():
            raise ValidationError({'branch': 'This branch already has user accounts and cannot be deleted.'})
        return super().destroy(request, *args, **kwargs)

class DynamicDischargeSummaryView(APIView):
    def _clean_status(self, raw_status):
        status_str = str(raw_status).upper()
        if "LAMA" in status_str: return "LAMA"
        if "DOPR" in status_str: return "DOPR"
        if "REFER" in status_str: return "REFER"
        if "DEATH" in status_str: return "DEATH"
        return "NORMAL"

    def get(self, request, uhid, adm_no):
        raw_type = request.query_params.get('type', 'NORMAL')
        summary_type = self._clean_status(raw_type)
        admission = get_object_or_404(Admission, patient__uhid=uhid, admNo=adm_no)

        existing_summary = DischargeSummary.objects.filter(admission=admission).first()
        if existing_summary:
            return Response({
                "is_existing": True,
                "summary_type": existing_summary.summary_type,
                "content": existing_summary.content
            }, status=status.HTTP_200_OK)

        template = copy.deepcopy(DISCHARGE_TEMPLATES.get(summary_type, DISCHARGE_TEMPLATES["NORMAL"]))

        # 🌟 UPDATED PRE-FILL: Iterate through the List to find keys
        try:
            med_hist = getattr(admission, 'medicalHistory', None)
            if med_hist:
                for section in template["sections"]:
                    if section["key"] == "k_c_o" and med_hist.previousDiagnosis:
                        section["value"] = med_hist.previousDiagnosis
        except Exception:
            pass

        return Response({"is_existing": False, "summary_type": summary_type, "content": template}, status=status.HTTP_200_OK)

    def post(self, request, uhid, adm_no):
        admission = get_object_or_404(Admission, patient__uhid=uhid, admNo=adm_no)
        raw_type = request.data.get('summary_type', 'NORMAL')
        summary_type = self._clean_status(raw_type)
        content = request.data.get('content', {})

        summary, created = DischargeSummary.objects.update_or_create(
            admission=admission,
            defaults={'summary_type': summary_type, 'content': content, 'created_by': request.user if request.user.is_authenticated else None}
        )
        return Response({"message": "Discharge Summary saved successfully!", "data": DischargeSummarySerializer(summary).data}, status=status.HTTP_200_OK)

class PrintDischargeSummaryView(APIView):
    def get(self, request, uhid, adm_no):
        admission = get_object_or_404(Admission, patient__uhid=uhid, admNo=adm_no)
        summary = DischargeSummary.objects.filter(admission=admission).first()
        if not summary:
            discharge = getattr(admission, 'discharge', None)
            raw_status = getattr(discharge, 'dischargeStatus', 'NORMAL')
            status_str = str(raw_status).upper()
            if "LAMA" in status_str:
                fallback_type = "LAMA"
            elif "DOPR" in status_str:
                fallback_type = "DOPR"
            elif "REFER" in status_str:
                fallback_type = "REFER"
            elif "DEATH" in status_str:
                fallback_type = "DEATH"
            else:
                fallback_type = "NORMAL"
            summary = DischargeSummary(
                admission=admission,
                summary_type=fallback_type,
                content=copy.deepcopy(DISCHARGE_TEMPLATES.get(fallback_type, DISCHARGE_TEMPLATES["NORMAL"])),
            )
        
        status_map = {"NORMAL": "pdf/normal.html", "RECOVERED": "pdf/normal.html", "LAMA": "pdf/lama.html", "REFER": "pdf/refer.html", "DOPR": "pdf/dopr.html", "DEATH": "pdf/death.html"}
        template_file = status_map.get(summary.summary_type, "pdf/normal.html")
        
        patient = admission.patient
        discharge = getattr(admission, 'discharge', None)
        billing = admission.bills.order_by('-id').first()

        age = "--"
        if patient.dob:
            calc_age = (timezone.now().date() - patient.dob).days // 365
            age = f"{calc_age} YRS"

        sections = summary.content.get("sections", [])

        # 🌟 NEW: BACKEND AUTO-CONVERTER 🌟
        # If an old database record is an Object/Dict, convert it to a List format instantly!
        if isinstance(sections, dict):
            sections = [{"key": k, **v} for k, v in sections.items()]

        # Now we can safely iterate through the list without crashing
        if discharge:
            for section in sections:
                if section.get("key") == "condition_at_discharge":
                    section["value"] = discharge.dischargeStatus.upper() if discharge.dischargeStatus else "--"

        context = {
            "s": summary, "sections": sections, "uhid": patient.uhid,
            "ipd_no": admission.ipdNo, "patient_name": patient.patientName.upper(),
            "guardian_name": patient.guardianName.upper() if patient.guardianName else "--",
            "address": patient.address, "consultant": discharge.doctorName.upper() if discharge and discharge.doctorName else "--",
            "claim_id": patient.tpaPanelCardNo if patient.tpaPanelCardNo else "--",
            "doa": admission.dateTime.strftime("%d-%m-%Y %H:%M HRS") if admission.dateTime else "--",
            "dod": discharge.dod.strftime("%d-%m-%Y %H:%M HRS") if (discharge and discharge.dod) else "--",
            "bill_no": f"{billing.id}/{admission.dateTime.strftime('%y')}" if billing else "--",
            "bill_date": discharge.dod.strftime("%d-%m-%Y %H:%M HRS") if (discharge and discharge.dod) else "--",
            "age_sex": f"{age} / {patient.gender.upper()}", "card_no": patient.tpaCard if patient.tpaCard else "--",
            "room": f"{discharge.roomNo} / {discharge.wardName.upper()}" if discharge and discharge.roomNo else "-- / --",
            "panel": patient.tpa.upper() if patient.tpa else patient.payMode.upper(),
            "contact_no": patient.phone, "status_on_discharge": discharge.dischargeStatus.upper() if discharge and discharge.dischargeStatus else "--",
        }

        html_string = render_to_string(template_file, context)
        result = io.BytesIO()
        pdf = pisa.pisaDocument(io.BytesIO(html_string.encode("UTF-8")), result)
        
        if not pdf.err:
            response = HttpResponse(result.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f'inline; filename="{uhid}_summary.pdf"'
            return response
            
        return Response({"error": "PDF Generation Failed"}, status=400)
    
class TaskReportAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if request.user.role not in ['superadmin', 'office_admin', 'admin', 'hod']:
            return Response({"error": "Unauthorized"}, status=status.HTTP_403_FORBIDDEN)

        staff = CustomUser.objects.filter(role__in=[
    'billing', 'opd', 'intimation', 'query', 'uploading', 'hod',
    'nursing', 'notes', 'medical_officer', 'quality_analyst',
])
        if request.user.role == 'admin':
            staff = staff.filter(branch=request.user.branch)
        
        report_data = []
        for employee in staff:
            total_tasks = employee.tasks_received.count()
            completed_tasks = employee.tasks_received.filter(status='Completed').count()
            
            assigned_patients = Patient.objects.filter(assigned_tasks__assigned_to=employee).distinct()
            patient_list = [{"uhid": p.uhid, "name": p.patientName} for p in assigned_patients]

            if total_tasks > 0:
                report_data.append({
                    "employee_name": f"{employee.first_name} {employee.last_name}",
                    "department": employee.role,
                    "total_tasks": total_tasks,
                    "completed_tasks": completed_tasks,
                    "pending_tasks": total_tasks - completed_tasks,
                    "assigned_patients": patient_list
                })

        return Response(report_data)
    
class TaskListCreateAPIView(generics.ListCreateAPIView):
    serializer_class = TaskSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return get_task_queryset_for_user(self.request.user)

    def perform_create(self, serializer):
        assigned_to = serializer.validated_data['assigned_to']
        patient = serializer.validated_data.get('patient')
        department = serializer.validated_data.get('department')
        validate_generic_task_assignment(
            self.request.user,
            assigned_to,
            patient=patient,
            department=department,
        )
        serializer.save(assigned_by=self.request.user)

class TaskDetailAPIView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = TaskSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return get_task_queryset_for_user(self.request.user)

    def perform_update(self, serializer):
        user = self.request.user

        # 🌟 THE BYPASS: If a regular employee is just clicking "Submit"
        if user.role not in TASK_MANAGER_ROLES:
            # Security check: Make sure they aren't trying to maliciously re-assign the task
            if 'assigned_to' in serializer.validated_data or 'department' in serializer.validated_data or 'patient' in serializer.validated_data:
                raise PermissionDenied("Employees cannot re-assign tasks or change departments.")
            
            # If they are just updating status (like "Completed") or adding notes, let it save!
            serializer.save()
            return

        # 👔 THE MANAGER CHECK: If an Admin/HOD is editing, run the strict validation
        assigned_to = serializer.validated_data.get('assigned_to', serializer.instance.assigned_to)
        patient = serializer.validated_data.get('patient', serializer.instance.patient)
        department = serializer.validated_data.get('department', serializer.instance.department)
        
        validate_generic_task_assignment(
            user,
            assigned_to,
            patient=patient,
            department=department,
        )
        serializer.save()

class LabReportListCreateView(generics.ListCreateAPIView):
    serializer_class = LabReportSerializer

    def get_queryset(self):
        uhid = self.kwargs.get('uhid')
        adm_no = self.kwargs.get('adm_no')
        return LabReport.objects.filter(patient__uhid=uhid, admission__admNo=adm_no).order_by('report_date', 'id')

    def perform_create(self, serializer):
        uhid = self.kwargs.get('uhid')
        adm_no = self.kwargs.get('adm_no')
        
        patient = get_object_or_404(Patient, uhid=uhid)
        admission = get_object_or_404(Admission, patient=patient, admNo=adm_no)

        lookup = {
            'patient': patient,
            'admission': admission,
            'report_name': serializer.validated_data.get('report_name'),
            'report_type': serializer.validated_data.get('report_type', ''),
            'report_date': serializer.validated_data.get('report_date'),
        }
        defaults = {
            **serializer.validated_data,
            'created_by': self.request.user.first_name or self.request.user.username,
        }
        report, _ = LabReport.objects.update_or_create(defaults=defaults, **lookup)
        serializer.instance = report

class LabReportBulkSaveAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, uhid, adm_no):
        patient = get_object_or_404(Patient, uhid=uhid)
        admission = get_object_or_404(Admission, patient=patient, admNo=adm_no)
        reports = request.data.get('reports') or []

        created_by = request.user.first_name or request.user.username
        created_reports = []

        with transaction.atomic():
            LabReport.objects.filter(patient=patient, admission=admission).delete()

            for report in reports:
                serializer = LabReportSerializer(data=report)
                serializer.is_valid(raise_exception=True)
                created_reports.append(LabReport(
                    patient=patient,
                    admission=admission,
                    created_by=created_by,
                    **serializer.validated_data,
                ))

            if created_reports:
                LabReport.objects.bulk_create(created_reports)

        payload = LabReportSerializer(
            LabReport.objects.filter(patient=patient, admission=admission).order_by('report_date', 'id'),
            many=True,
        ).data
        return Response(payload, status=status.HTTP_200_OK)


class LabReportTemplateSuggestionsAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, uhid, adm_no):
        patient = get_object_or_404(Patient, uhid=uhid)
        admission = get_object_or_404(Admission, patient=patient, admNo=adm_no)

        suggested_reports = build_suggested_reports_for_admission(patient, admission)
        return Response(
            {
                'patient': patient.uhid,
                'admNo': admission.admNo,
                'suggested_reports': suggested_reports,
            },
            status=status.HTTP_200_OK,
        )

class HODEmployeeListAPIView(APIView):
    def get(self, request):
        denied = ensure_hod_access(request)
        if denied:
            return denied

        department = request.query_params.get('department')
        role_slug = get_department_role(department)
        if not role_slug:
            return Response({'error': 'Invalid department.'}, status=status.HTTP_400_BAD_REQUEST)
        queryset = CustomUser.objects.filter(role=role_slug)

        if getattr(request.user, 'branch', None) in get_valid_branch_codes():
            queryset = queryset.filter(branch=request.user.branch)

        queryset = (queryset | CustomUser.objects.filter(pk=request.user.pk)).distinct()

        employees = []
        for employee in queryset.order_by('first_name', 'username'):
            tasks = employee.tasks_received.filter(department=department)
            employee_name = employee.get_full_name().strip() or employee.username
            employees.append({
                'id': employee.id,
                'employeeCode': employee.emp_id or employee.username,
                'name': employee_name,
                'email': employee.email,
                'role': employee.role,
                'department': department,
                'taskCount': tasks.count(),
            })

        return Response({'employees': employees}, status=status.HTTP_200_OK)

class HODTaskListCreateAPIView(APIView):
    def get(self, request):
        denied = ensure_hod_access(request)
        if denied:
            return denied

        department = request.query_params.get('department')
        role_slug = get_department_role(department)
        if not role_slug:
            return Response({'error': 'Invalid department.'}, status=status.HTTP_400_BAD_REQUEST)
        employee_id = request.query_params.get('employeeId')
        date_filter = request.query_params.get('date')
        status_filter = request.query_params.get('status')

        tasks = Task.objects.select_related('assigned_to', 'patient').filter(
            models.Q(department=department) | models.Q(assigned_to=request.user)
        )

        if getattr(request.user, 'branch', None) in get_valid_branch_codes():
            tasks = tasks.filter(
                models.Q(assigned_to__branch=request.user.branch) |
                models.Q(assigned_to__branch__isnull=True)
            )

        if employee_id:
            tasks = tasks.filter(assigned_to_id=employee_id)
        if date_filter:
            tasks = tasks.filter(created_at__date=date_filter)

        serialized = [serialize_task_for_hod(task) for task in tasks.order_by('-created_at')]
        if status_filter:
            serialized = [task for task in serialized if task['status'] == status_filter]

        return Response({'tasks': serialized}, status=status.HTTP_200_OK)

    def post(self, request):
        denied = ensure_hod_access(request)
        if denied:
            return denied

        employee_id = request.data.get('employeeId')
        department = request.data.get('department')
        role_slug = get_department_role(department)
        if not role_slug:
            return Response({'error': 'Invalid department.'}, status=status.HTTP_400_BAD_REQUEST)
        assigned_to = get_object_or_404(CustomUser, pk=employee_id)
        if assigned_to.pk != request.user.pk and assigned_to.role != role_slug:
            return Response({'error': 'Selected employee does not belong to this department.'}, status=status.HTTP_400_BAD_REQUEST)
        if getattr(request.user, 'branch', None) in get_valid_branch_codes() and assigned_to.branch != request.user.branch:
            return Response({'error': 'You can assign tasks only inside your own branch.'}, status=status.HTTP_403_FORBIDDEN)
        due_date_raw = request.data.get('dueDate')
        due_date = None
        if due_date_raw:
            due_date = timezone.make_aware(datetime.datetime.fromisoformat(f"{due_date_raw}T23:59:00"))

        patient = None
        patient_uhid = request.data.get('patientId')
        if patient_uhid:
            patient = Patient.objects.filter(uhid=patient_uhid).first()
            if not patient:
                return Response({'error': 'Selected patient was not found.'}, status=status.HTTP_400_BAD_REQUEST)
            if assigned_to.branch in get_valid_branch_codes() and patient.branch_location != assigned_to.branch:
                return Response({'error': 'Selected patient belongs to a different branch.'}, status=status.HTTP_400_BAD_REQUEST)

        priority_map = {'low': 'Low', 'medium': 'Medium', 'high': 'High'}
        task = Task.objects.create(
            title=request.data.get('taskType') or 'Task',
            description=request.data.get('notes') or '',
            assigned_by=request.user,
            assigned_to=assigned_to,
            department=department,
            priority=priority_map.get(str(request.data.get('priority')).lower(), 'Medium'),
            status=normalize_task_status(request.data.get('status') or 'pending', due_date),
            due_date=due_date,
            patient=patient,
        )

        return Response({'task': serialize_task_for_hod(task)}, status=status.HTTP_201_CREATED)

class HODTaskDetailAPIView(APIView):
    def patch(self, request, pk):
        denied = ensure_hod_access(request)
        if denied:
            return denied

        task = get_object_or_404(Task, pk=pk)
        due_date = task.due_date
        if 'priority' in request.data:
            priority_map = {'low': 'Low', 'medium': 'Medium', 'high': 'High'}
            task.priority = priority_map.get(str(request.data.get('priority')).lower(), task.priority)
        if 'notes' in request.data:
            task.description = request.data.get('notes') or ''
        if 'status' in request.data:
            task.status = normalize_task_status(request.data.get('status'), due_date)
        task.save()
        return Response({'task': serialize_task_for_hod(task)}, status=status.HTTP_200_OK)

class HODAnalyticsAPIView(APIView):
    def get(self, request):
        denied = ensure_hod_access(request)
        if denied:
            return denied

        department = request.query_params.get('department')
        role_slug = get_department_role(department)
        if not role_slug:
            return Response({'error': 'Invalid department.'}, status=status.HTTP_400_BAD_REQUEST)
        employee_id = request.query_params.get('employeeId')
        date_filter = request.query_params.get('date')

        tasks = Task.objects.select_related('assigned_to', 'patient').filter(department=department)
        if getattr(request.user, 'branch', None) in get_valid_branch_codes():
            tasks = tasks.filter(assigned_to__branch=request.user.branch)
        if employee_id:
            tasks = tasks.filter(assigned_to_id=employee_id)
        if date_filter:
            tasks = tasks.filter(created_at__date=date_filter)

        task_rows = [serialize_task_for_hod(task) for task in tasks]
        total = len(task_rows)
        completed = sum(1 for task in task_rows if task['status'] == 'completed')
        pending = sum(1 for task in task_rows if task['status'] == 'pending')
        overdue = sum(1 for task in task_rows if task['status'] == 'overdue')

        stats = [
            {'label': 'Total Tasks', 'value': total, 'sub': f'{department or "All"} department'},
            {'label': 'Completed', 'value': completed, 'sub': 'Marked complete'},
            {'label': 'Pending', 'value': pending, 'sub': 'Awaiting action'},
            {'label': 'Overdue', 'value': overdue, 'sub': 'Past due date'},
        ]

        employee_stats = []
        employee_queryset = CustomUser.objects.filter(role=role_slug)
        if getattr(request.user, 'branch', None) in get_valid_branch_codes():
            employee_queryset = employee_queryset.filter(branch=request.user.branch)
        for employee in employee_queryset:
            employee_tasks = [task for task in task_rows if task['employeeId'] == employee.id]
            if not employee_tasks and employee_id:
                continue
            assigned = len(employee_tasks)
            emp_completed = sum(1 for task in employee_tasks if task['status'] == 'completed')
            emp_pending = sum(1 for task in employee_tasks if task['status'] == 'pending')
            emp_overdue = sum(1 for task in employee_tasks if task['status'] == 'overdue')
            employee_stats.append({
                'id': employee.id,
                'name': employee.get_full_name().strip() or employee.username,
                'assigned': assigned,
                'completed': emp_completed,
                'pending': emp_pending,
                'overdue': emp_overdue,
                'completionPct': int(round((emp_completed / assigned) * 100)) if assigned else 0,
            })

        return Response({'stats': stats, 'employeeStats': employee_stats}, status=status.HTTP_200_OK)

class HODReviewListCreateAPIView(APIView):
    def get(self, request):
        denied = ensure_hod_access(request)
        if denied:
            return denied

        department = request.query_params.get('department')
        reviews = HODReview.objects.select_related('employee', 'reviewed_by').filter(department=department).order_by('-created_at')

        if getattr(request.user, 'branch', None) in get_valid_branch_codes():
            reviews = reviews.filter(employee__branch=request.user.branch)

        payload = []
        for review in reviews:
            payload.append({
                'id': review.id,
                'employeeName': review.employee.get_full_name().strip() or review.employee.username,
                'employeeId': review.employee_id,
                'period': review.period,
                'rating': review.rating,
                'performanceScore': review.performance_score,
                'comments': review.comments,
                'submittedAt': timezone.localtime(review.created_at).strftime('%Y-%m-%d %H:%M'),
            })
        return Response({'reviews': payload}, status=status.HTTP_200_OK)

    def post(self, request):
        denied = ensure_hod_access(request)
        if denied:
            return denied

        employee = get_object_or_404(CustomUser, pk=request.data.get('employeeId'))
        review = HODReview.objects.create(
            department=request.data.get('department') or '',
            employee=employee,
            reviewed_by=request.user,
            period=request.data.get('period') or 'weekly',
            rating=int(request.data.get('rating') or 5),
            performance_score=str(request.data.get('performanceScore') or ''),
            comments=request.data.get('comments') or '',
            task_name=request.data.get('taskName') or 'Department Performance',
        )
        serializer = HODReviewSerializer(review)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

class HODReportDownloadAPIView(APIView):
    def get(self, request):
        denied = ensure_hod_access(request)
        if denied:
            return denied

        department = request.query_params.get('department')
        role_slug = get_department_role(department)
        if not role_slug:
            return Response({'error': 'Invalid department.'}, status=status.HTTP_400_BAD_REQUEST)
        employee_id = request.query_params.get('employeeId')
        date_filter = request.query_params.get('date')

        tasks = Task.objects.select_related('assigned_to', 'patient').filter(department=department)
        if getattr(request.user, 'branch', None) in get_valid_branch_codes():
            tasks = tasks.filter(assigned_to__branch=request.user.branch)
        if employee_id:
            tasks = tasks.filter(assigned_to_id=employee_id)
        if date_filter:
            tasks = tasks.filter(created_at__date=date_filter)

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{quote((department or "department").lower())}_tasks_report.csv"'
        writer = csv.writer(response)
        writer.writerow(['Task ID', 'Employee', 'Task', 'Patient UHID', 'Priority', 'Status', 'Due Date'])
        for task in tasks.order_by('-created_at'):
            data = serialize_task_for_hod(task)
            writer.writerow([
                data['id'],
                data['employeeName'],
                data['taskType'],
                data['patientId'],
                data['priority'],
                data['status'],
                data['dueDate'],
            ])
        return response

class PerformanceRatingsAPIView(APIView):
    def get(self, request):
        if not request.user.is_authenticated:
            return Response({'error': 'Unauthorized access.'}, status=status.HTTP_401_UNAUTHORIZED)

        reviews = HODReview.objects.select_related('employee', 'reviewed_by').order_by('-created_at')
        payload = []
        for review in reviews:
            branch_obj = HospitalSettings.objects.filter(branch=review.employee.branch).first()
            branch = branch_obj.slug if branch_obj else str(review.employee.branch or '').lower()
            payload.append({
                'staffName': review.employee.get_full_name().strip() or review.employee.username,
                'staffId': review.employee.emp_id or review.employee.username,
                'branch': branch,
                'role': review.employee.get_role_display(),
                'department': review.department,
                'task': review.task_name or 'Department Performance',
                'rating': review.rating,
                'reviewedBy': review.reviewed_by.get_full_name().strip() if review.reviewed_by else 'System',
                'description': review.comments,
                'reason': '',
                'date': timezone.localtime(review.created_at).date().isoformat(),
            })
        return Response(payload, status=status.HTTP_200_OK)

class DepartmentLogListAPIView(APIView):
    def get(self, request):
        if not request.user.is_authenticated:
            return Response({'error': 'Unauthorized access.'}, status=status.HTTP_401_UNAUTHORIZED)

        department = request.query_params.get('department')
        queryset = DepartmentLogEntry.objects.filter(department=department)

        if getattr(request.user, 'branch', None) in get_valid_branch_codes():
            queryset = queryset.filter(branch=request.user.branch)

        serializer = DepartmentLogEntrySerializer(queryset.order_by('-record_date', '-created_at'), many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

class DepartmentLogBulkSaveAPIView(APIView):
    def post(self, request):
        if not request.user.is_authenticated:
            return Response({'error': 'Unauthorized access.'}, status=status.HTTP_401_UNAUTHORIZED)

        department = request.data.get('department')
        entries = request.data.get('entries') or []
        if department not in dict(DepartmentLogEntry.DEPARTMENT_CHOICES):
            return Response({'error': 'Invalid department.'}, status=status.HTTP_400_BAD_REQUEST)

        branch = getattr(request.user, 'branch', None) if getattr(request.user, 'branch', None) in get_valid_branch_codes() else resolve_branch_code_from_loc(None, request.data.get('branch'))
        record_dates = sorted({coerce_record_date(department, entry) for entry in entries})

        with transaction.atomic():
            if record_dates:
                DepartmentLogEntry.objects.filter(
                    department=department,
                    branch=branch,
                    record_date__in=record_dates,
                ).delete()

            created = []
            for entry in entries:
                created.append(DepartmentLogEntry(
                    department=department,
                    branch=branch,
                    record_date=coerce_record_date(department, entry),
                    data=entry,
                    created_by=request.user,
                ))
            if created:
                DepartmentLogEntry.objects.bulk_create(created)

        return Response({'saved': len(entries)}, status=status.HTTP_200_OK)
    
class ReportMasterViewSet(viewsets.ModelViewSet):
    queryset = ReportMaster.objects.all().order_by('name')
    serializer_class = ReportMasterSerializer
    permission_classes = [IsAuthenticated]

class MedicineMasterViewSet(viewsets.ModelViewSet):
    queryset = MedicineMaster.objects.all().order_by('name')
    serializer_class = MedicineMasterSerializer
    permission_classes = [IsAuthenticated]


def parse_medicine_master_workbook(uploaded_file):
    def normalize_expiry_date(value):
        if value in (None, ''):
            return None

        if isinstance(value, datetime.datetime):
            return value.date()
        if isinstance(value, datetime.date):
            return value

        text = str(value).strip()
        if not text:
            return None

        for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y", "%d-%m-%y"):
            try:
                return datetime.datetime.strptime(text, fmt).date()
            except ValueError:
                continue

        for fmt in ("%m/%Y", "%m-%Y", "%m/%y", "%m-%y"):
            try:
                parsed = datetime.datetime.strptime(text, fmt)
                return datetime.date(parsed.year, parsed.month, 1)
            except ValueError:
                continue

        return None

    workbook = openpyxl.load_workbook(uploaded_file, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    header_row_index = None
    headers = []
    for index, row in enumerate(worksheet.iter_rows(min_row=1, max_row=min(25, worksheet.max_row), values_only=True), start=1):
        normalized = [str(cell).strip().lower() if cell is not None else '' for cell in row]
        if 'description' in normalized and 'rate' in normalized and ('qty.' in normalized or 'qty' in normalized):
            header_row_index = index
            headers = [str(cell).strip() if cell is not None else '' for cell in row]
            break

    if not header_row_index:
        raise ValidationError({'file': "Could not find medicine sheet headers. Expected columns like Description, Batch No., Exp., Rate, Qty."})

    parsed_rows = []
    for row in worksheet.iter_rows(min_row=header_row_index + 1, values_only=True):
        row_map = {headers[idx]: row[idx] if idx < len(row) else None for idx in range(len(headers))}

        description = str(row_map.get('Description') or '').strip()
        if not description or description.lower() in {'none', 'nan'}:
            continue

        batch_no = str(row_map.get('Batch No.') or '').strip()
        expiry_date = normalize_expiry_date(row_map.get('Exp.'))
        rate_raw = row_map.get('Rate')
        qty_raw = row_map.get('Qty.') if 'Qty.' in row_map else row_map.get('Qty')

        try:
            rate = Decimal(str(rate_raw or 0)).quantize(Decimal('0.01'))
        except Exception:
            rate = Decimal('0.00')

        try:
            quantity = int(float(qty_raw or 0))
        except Exception:
            quantity = 0

        parsed_rows.append(
            MedicineMaster(
                name=description,
                batch_no=batch_no or None,
                expiry_date=expiry_date,
                rate=rate,
                quantity=max(quantity, 0),
            )
        )

    if not parsed_rows:
        raise ValidationError({'file': 'No medicine rows were found in the uploaded sheet.'})

    return parsed_rows


class MedicineMasterImportAPIView(APIView):
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        if getattr(request.user, 'role', '') not in {'superadmin', 'office_admin'}:
            raise PermissionDenied("Only Super Admin and Office Admin can import medicine records.")

        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            raise ValidationError({'file': 'Please upload an Excel file.'})

        rows = parse_medicine_master_workbook(uploaded_file)

        with transaction.atomic():
            MedicineMaster.objects.all().delete()
            MedicineMaster.objects.bulk_create(rows)

        return Response(
            {
                'imported': len(rows),
                'message': 'Medicine master updated successfully.',
                'sample': MedicineMasterSerializer(MedicineMaster.objects.all().order_by('name')[:10], many=True).data,
            },
            status=status.HTTP_200_OK,
        )

class PharmacyRecordViewSet(viewsets.ModelViewSet):
    serializer_class = PharmacyRecordSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        # Using exact kwarg names 'uhid' and 'adm_no' from urls.py
        return PharmacyRecord.objects.filter(
            patient__uhid=self.kwargs['uhid'],
            admission__admNo=self.kwargs['adm_no']
        )

    def perform_create(self, serializer):
        patient = get_object_or_404(Patient, uhid=self.kwargs['uhid'])
        admission = get_object_or_404(Admission, admNo=self.kwargs['adm_no'], patient=patient)
        lookup = {
            'patient': patient,
            'admission': admission,
            'medicine_name': serializer.validated_data.get('medicine_name'),
            'date_given': serializer.validated_data.get('date_given'),
        }
        defaults = {
            **serializer.validated_data,
            'created_by': self.request.user,
        }
        record, _ = PharmacyRecord.objects.update_or_create(defaults=defaults, **lookup)
        serializer.instance = record

class PharmacyRecordBulkSaveAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, uhid, adm_no):
        patient = get_object_or_404(Patient, uhid=uhid)
        admission = get_object_or_404(Admission, admNo=adm_no, patient=patient)
        records = request.data.get('records') or []

        created_records = []
        with transaction.atomic():
            PharmacyRecord.objects.filter(patient=patient, admission=admission).delete()

            for record in records:
                serializer = PharmacyRecordSerializer(data=record)
                serializer.is_valid(raise_exception=True)
                created_records.append(PharmacyRecord(
                    patient=patient,
                    admission=admission,
                    created_by=request.user,
                    **serializer.validated_data,
                ))

            if created_records:
                PharmacyRecord.objects.bulk_create(created_records)

        payload = PharmacyRecordSerializer(
            PharmacyRecord.objects.filter(patient=patient, admission=admission).order_by('date_given', 'id'),
            many=True,
        ).data
        return Response(payload, status=status.HTTP_200_OK)

class TaskEligibleEmployeesAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        department = request.query_params.get('department')

        if user.role == 'superadmin':
            employees = CustomUser.objects.filter(role__in=TASK_ASSIGNABLE_ROLES | {'admin', 'office_admin'})
        elif user.role == 'office_admin':
            employees = CustomUser.objects.filter(role__in=TASK_ASSIGNABLE_ROLES)
        elif user.role == 'admin':
            employees = CustomUser.objects.filter(role__in=TASK_ASSIGNABLE_ROLES, branch=user.branch)
        elif user.role == 'hod':
            role_slug = get_department_role(department)
            if not role_slug:
                return Response({"error": "Invalid department."}, status=status.HTTP_400_BAD_REQUEST)
            employees = CustomUser.objects.filter(role=role_slug)
        if getattr(user, 'branch', None) in get_valid_branch_codes():
                employees = employees.filter(branch=user.branch)
        else:
            employees = CustomUser.objects.none()

        data = [{"id": emp.id, "name": emp.get_full_name().strip() or emp.username, "role": emp.get_role_display()} for emp in employees]
        return Response(data)

class BulkTaskAssignAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        # 1. Check if user is in the updated TASK_MANAGER_ROLES
        if request.user.role not in TASK_MANAGER_ROLES:
            return Response({"error": "Unauthorized"}, status=status.HTTP_403_FORBIDDEN)

        serializer = BulkTaskAssignSerializer(data=request.data)
        if serializer.is_valid():
            assign_to_id = serializer.validated_data['assign_to']
            patient_ids = serializer.validated_data['patient_ids']
            department = serializer.validated_data['department']
            title = serializer.validated_data.get('title', 'Patient Task')

            try:
                assigned_to_user = CustomUser.objects.get(id=assign_to_id)
            except CustomUser.DoesNotExist:
                return Response({"error": "Assigned user not found."}, status=status.HTTP_404_NOT_FOUND)

            try:
                validate_generic_task_assignment(request.user, assigned_to_user)
            except PermissionDenied as exc:
                return Response({"error": str(exc.detail)}, status=status.HTTP_403_FORBIDDEN)
            except ValidationError as exc:
                return Response(exc.detail, status=status.HTTP_400_BAD_REQUEST)

            tasks_to_create = []
            for pid in patient_ids:
                try:
                    patient = Patient.objects.get(id=pid)
                    

                    validate_generic_task_assignment(
                        request.user,
                        assigned_to_user,
                        patient=patient,
                    )
                    
                    tasks_to_create.append(
                        Task(
                            title=title,
                            assigned_by=request.user,
                            assigned_to=assigned_to_user,
                            department=department,
                            patient=patient,
                            status='Pending'
                        )
                    )
                except ValidationError as exc:
                    return Response(exc.detail, status=status.HTTP_400_BAD_REQUEST)
                except Patient.DoesNotExist:
                    continue

            Task.objects.bulk_create(tasks_to_create)
            
            return Response(
                {"message": f"Successfully assigned {len(tasks_to_create)} patients to {assigned_to_user.username}."}, 
                status=status.HTTP_201_CREATED
            )
            
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class TaskAnalyticsAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        
        # 👇 FIXED: Using exact database role keys (lowercase)
        if user.role not in ['superadmin', 'office_admin', 'hod']:
            return Response({"error": "Unauthorized"}, status=status.HTTP_403_FORBIDDEN)

        if user.role == 'hod':
            qs = Task.objects.filter(assigned_by=user)
        else:
            qs = Task.objects.all()

        # 👇 FIXED: Changed assigned_to__name to assigned_to__username
        analytics = qs.values(
            'assigned_to__id', 
            'assigned_to__username', 
            'assigned_to__role'
        ).annotate(
            total_tasks=Count('id'),
            completed_tasks=Count('id', filter=Q(status='Completed')),
            pending_tasks=Count('id', filter=Q(status='Pending'))
        )
        return Response(analytics)
    
class EmployeeTaskUpdateAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def patch(self, request, task_id):
        try:
            task = Task.objects.get(id=task_id)
        except Task.DoesNotExist:
            return Response({"error": "Task not found."}, status=status.HTTP_404_NOT_FOUND)
        
        user = request.user
        is_admin_or_hod = user.role in ['superadmin', 'office_admin', 'hod']

        # 1. Authorization: Must be the assigned employee OR an admin/hod
        if task.assigned_to != user and not is_admin_or_hod:
            return Response({"error": "Not authorized to update this task."}, status=status.HTTP_403_FORBIDDEN)

        # 2. 🌟 THE LOCK: Block employee if completed, but let Admin/HOD bypass
        if task.status == 'Completed' and not is_admin_or_hod:
            return Response(
                {"error": "This task is already submitted and locked. Only an HOD or Admin can edit it now."}, 
                status=status.HTTP_403_FORBIDDEN
            )
        
        # 3. Normalize Status
        raw_status = str(request.data.get('status', task.status)).strip().title()
        if raw_status.lower() == 'completed':
            raw_status = 'Completed'
        elif raw_status.lower() in ['in progress', 'inprogress']:
            raw_status = 'In Progress'

        valid_statuses = ['Pending', 'In Progress', 'Completed', 'On Hold', 'Overdue']
        
        if raw_status in valid_statuses:
            task.status = raw_status
            
            # 4. Capture the work & label who wrote it
            work_done = request.data.get('work_done') or request.data.get('remarks') or request.data.get('notes')
            if work_done:
                role_label = "HOD/Admin" if is_admin_or_hod else "Employee"
                if task.description:
                    task.description = f"{task.description}\n\n[{role_label} Update]: {work_done}"
                else:
                    task.description = f"[{role_label} Update]: {work_done}"

            task.save()
            return Response({
                "message": "Task updated successfully!", 
                "status": task.status,
                "notes": task.description
            }, status=status.HTTP_200_OK)
            
        return Response({"error": f"Invalid status. Must be one of {valid_statuses}"}, status=status.HTTP_400_BAD_REQUEST)

class EmployeeMyTasksAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        tasks = get_task_queryset_for_user(request.user)
        
        # Sort logic: 
        # 1. Pending (Value 1) comes first, then everything else (Value 2)
        # 2. Older tasks come first ('created_at' ascending)
        tasks = tasks.annotate(
    status_order=Case(
        When(status='Pending', then=Value(1)),
        When(status='In Progress', then=Value(2)), 
        When(status='Overdue', then=Value(3)),
        default=Value(4),                             
        output_field=IntegerField(),
    )
).order_by('status_order', 'created_at')

        serializer = TaskSerializer(tasks, many=True, context={'request': request})
        return Response(serializer.data)
class DoctorViewSet(viewsets.ModelViewSet):
    queryset = Doctor.objects.all().order_by('name')
    serializer_class = DoctorSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = None # Return all doctors at once for the dropdown

    # Restrict POST/PUT/DELETE to Admins only
    def check_permissions(self, request):
        super().check_permissions(request)
        if request.method not in ['GET', 'HEAD', 'OPTIONS']:
            if getattr(request.user, 'role', '') not in ['superadmin', 'admin', 'office_admin']:
                self.permission_denied(request, message="Only Admins can manage the doctors list.")


class PrintBillView(APIView):
    # If you want to test this easily in the browser without a token, uncomment the line below temporarily:
    permission_classes = [] 

    def get(self, request, uhid, adm_no):
        admission = get_object_or_404(Admission, patient__uhid=uhid, admNo=adm_no)
        patient = admission.patient
        discharge = getattr(admission, 'discharge', None)
        
        billing_obj, _ = get_or_create_current_billing(admission)
        services = admission.services.all().order_by('svcDate', 'id')

        # 🧮 1. Calculate Totals
        gross_total = sum((svc.svcTot or 0) for svc in services)
        discount = billing_obj.discount or Decimal('0.00')
        advance = billing_obj.advance or Decimal('0.00')
        net_payable = gross_total - discount - advance

        age = "--"
        if patient.dob:
            calc_age = (timezone.now().date() - patient.dob).days // 365
            age = f"{calc_age} YRS"

        # 🌟 2. Fetch Dynamic Hospital Settings based on the PATIENT'S BRANCH!
        settings_obj = HospitalSettings.objects.filter(branch=patient.branch_location).first()
        
        # Fallback just in case the Admin hasn't created settings for this branch yet
        if not settings_obj:
            settings_obj = HospitalSettings.objects.first()

        logo_base64 = ""
        
        # First, try to use the logo uploaded via the Admin panel
        if settings_obj and settings_obj.logo and hasattr(settings_obj.logo, 'path'):
            if os.path.exists(settings_obj.logo.path):
                with open(settings_obj.logo.path, "rb") as image_file:
                    logo_base64 = base64.b64encode(image_file.read()).decode("utf-8")
        
        # If no custom logo is uploaded, fallback to the default static logo
        if not logo_base64:
            logo_path = os.path.join(settings.BASE_DIR, 'static', 'logo.png')
            if os.path.exists(logo_path):
                with open(logo_path, "rb") as image_file:
                    logo_base64 = base64.b64encode(image_file.read()).decode("utf-8")
        # 🌟 3. Generate the QR Code dynamically (using the dynamic website URL!)
        qr_url = settings_obj.website if settings_obj and settings_obj.website else "https://sangihospital.com/"
        qr = qrcode.make(qr_url)
        buffer = io.BytesIO()
        qr.save(buffer, format="PNG")
        qr_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")

        # 🌟 4. Safely handle Room/Ward to prevent NoneType crash if patient isn't discharged
        safe_room = discharge.roomNo if discharge and discharge.roomNo else '--'
        safe_ward = discharge.wardName.upper() if discharge and discharge.wardName else '--'
        room_ward = f"{safe_room} / {safe_ward}"

        # 📋 5. Build the context for the HTML template
        context = {
            "current_date": timezone.now().strftime("%d/%m/%Y"),
            "admission_type": admission.admissionType.upper(),
            "uhid": patient.uhid,
            "bill_no": f"{billing_obj.id}/{admission.dateTime.strftime('%y')}" if billing_obj else "--",
            "ipd_no": admission.ipdNo or "--",
            "bill_date": timezone.now().strftime("%d/%m/%Y %H:%M HRS"),
            "patient_name": patient.patientName.upper(),
            "age_sex": f"{age} / {patient.gender.upper()}",
            "guardian_name": patient.guardianName.upper() if patient.guardianName else "--",
            "address": patient.address or "--",
            "consultant": discharge.doctorName.upper() if discharge and discharge.doctorName else "--",
            "room_ward": room_ward,
            "claim_id": patient.tpaPanelCardNo or "--",
            "panel": patient.tpa.upper() if patient.tpa else patient.payMode.upper(),
            "doa": timezone.localtime(admission.dateTime).strftime("%d/%m/%Y, %I:%M %p") if admission.dateTime else "--",
            "contact_no": patient.phone or "--",
            "dod": timezone.localtime(discharge.dod).strftime("%d/%m/%Y, %I:%M %p") if discharge and discharge.dod else "--",
            "discharge_status": discharge.dischargeStatus.upper() if discharge and discharge.dischargeStatus else "--",
            
            # The Data & Math
            "services": services,
            "gross_total": f"{gross_total:,.2f}",
            "discount": f"{discount:,.2f}",
            "advance": f"{advance:,.2f}",
            "net_payable": f"{net_payable:,.2f}",
            
            # The Images
            "qr_code": qr_base64,
            "logo_base64": logo_base64,
            "hospital": settings_obj, 
        }

        # 🖨️ 6. Render the PDF
        html_string = render_to_string("pdf/bill.html", context)
        result = io.BytesIO()
        pdf = pisa.pisaDocument(io.BytesIO(html_string.encode("UTF-8")), result)
        
        if not pdf.err:
            response = HttpResponse(result.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f'inline; filename="{patient.uhid}_final_bill.pdf"'
            return response
            
        return Response({"error": "PDF Generation Failed"}, status=status.HTTP_400_BAD_REQUEST)

class AdminDashboardStatsAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        
        # 1. Security: Only Admins can see this dashboard data
        if user.role not in ['superadmin', 'office_admin', 'admin']:
            return Response({"error": "Unauthorized"}, status=status.HTTP_403_FORBIDDEN)

        today = timezone.now().date()
        
        # 2. Get all discharges where the Date of Discharge (dod) is today
        todays_discharges = Discharge.objects.filter(dod__date=today)

        # 3. If it's a Branch Admin, strictly filter to ONLY show their branch!
        if user.role == 'admin':
            todays_discharges = todays_discharges.filter(admission__patient__branch_location=user.branch)

        # 4. Return the exact count to the frontend
        return Response({
            "todaysDischargeCount": todays_discharges.count(),
            # 💡 Pro-tip: You can easily add more stats here later! 
            # Example: "totalPatients": Patient.objects.filter(branch_location=user.branch).count()
        }, status=status.HTTP_200_OK)

class PrintAdmissionNoteView(APIView):
    # Remove this line once you add token auth to the frontend call:
    permission_classes = []
 
    def get(self, request, uhid, adm_no):
        # ── 1. Fetch core objects ──────────────────────────────────────────
        admission  = get_object_or_404(Admission, patient__uhid=uhid, admNo=adm_no)
        patient    = admission.patient
        discharge  = getattr(admission, 'discharge', None)
        med_hist   = getattr(admission, 'medicalHistory', None)
 
        # ── 2. Calculate age ──────────────────────────────────────────────
        age = "--"
        if patient.dob:
            age = f"{(timezone.now().date() - patient.dob).days // 365} YRS"
 
        # ── 3. Hospital settings + logo (same pattern as PrintBillView) ───
        settings_obj = HospitalSettings.objects.filter(branch=patient.branch_location).first()
        if not settings_obj:
            settings_obj = HospitalSettings.objects.first()
 
        logo_base64 = ""
        # Try uploaded logo first
        if settings_obj and settings_obj.logo and hasattr(settings_obj.logo, 'path'):
            if os.path.exists(settings_obj.logo.path):
                with open(settings_obj.logo.path, "rb") as f:
                    logo_base64 = base64.b64encode(f.read()).decode("utf-8")
        # Fallback to static/logo.png
        if not logo_base64:
            logo_path = os.path.join(settings.BASE_DIR, 'static', 'logo.png')
            if os.path.exists(logo_path):
                with open(logo_path, "rb") as f:
                    logo_base64 = base64.b64encode(f.read()).decode("utf-8")
 
        # ── 4. Ward / Bed ─────────────────────────────────────────────────
        ward_parts = []
        if discharge and discharge.wardName:
            ward_parts.append(discharge.wardName.upper())
        if discharge and discharge.bedNo:
            ward_parts.append(discharge.bedNo)
        ward_bed = " / ".join(ward_parts) if ward_parts else "--"
 
        # ── 5. Card no (TPA card or payMode) ─────────────────────────────
        card_no = patient.tpaCard or patient.tpaPanelCardNo or "--"
 
        # ── 6. Past history — combine previousDiagnosis + pastSurgeries ──
        past_parts = []
        if med_hist:
            if med_hist.previousDiagnosis:
                past_parts.append(med_hist.previousDiagnosis.strip())
            if med_hist.pastSurgeries:
                past_parts.append(med_hist.pastSurgeries.strip())
        past_history = "\n".join(past_parts) if past_parts else ""
 
        # ── 7. Build template context ─────────────────────────────────────
        context = {
            # Hospital
            "hospital":       settings_obj,
            "logo_base64":    logo_base64,
 
            # Patient basics
            "patient_name":   patient.patientName.upper(),
            "age_sex":        f"{age} / {patient.gender.upper()}",
            "ipd_no":         admission.ipdNo or "--",
            "card_no":        card_no,
            "ward_bed":       ward_bed,
            "admission_date": timezone.localtime(admission.dateTime).strftime("%d/%m/%Y AT %I:%M HRS")
                              if admission.dateTime else "--",
 
            # Medical history fields
            "present_complaints":  med_hist.presentComplaints  if med_hist else "",
            "chief_complaints":    med_hist.chiefComplaints     if med_hist else "",
            "investigations":      med_hist.investigations      if med_hist else "",
            "past_history":        past_history,
            "treatment_advised":   med_hist.treatmentAdvised    if med_hist else "",
 
            # Vitals
            "bp":   med_hist.bp    if med_hist else "",
            "pr":   med_hist.pr or (med_hist.pulse if med_hist else ""),
            "spo2": med_hist.spo2  if med_hist else "",
            "temp": med_hist.temp  if med_hist else "",
            "chest":med_hist.chest if med_hist else "",
            "cvs":  med_hist.cvs   if med_hist else "",
            "cns":  med_hist.cns   if med_hist else "",
            "pa":   med_hist.pa    if med_hist else "",
 
            # Diagnosis & doctor
            "provisional_diagnosis": med_hist.provisionalDiagnosis if med_hist else "",
            "treating_doctor":       (
                med_hist.treatingDoctor if (med_hist and med_hist.treatingDoctor)
                else (discharge.doctorName if discharge and discharge.doctorName else "--")
            ),
        }
 
        # ── 8. Render → PDF ───────────────────────────────────────────────
        html_string = render_to_string("pdf/admission_note.html", context)
        result = io.BytesIO()
        pdf = pisa.pisaDocument(io.BytesIO(html_string.encode("UTF-8")), result)
 
        if not pdf.err:
            response = HttpResponse(result.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = (
                f'inline; filename="{patient.uhid}_adm{adm_no}_admission_note.pdf"'
            )
            return response
 
        return Response({"error": "PDF Generation Failed"}, status=400)