import os
import django

# Setup Django environment
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'sangi_hospital.settings')
django.setup()

from patients.models import ServiceMaster

def run_import(filepath, is_cashless=False):
    try:
        import openpyxl
    except ImportError:
        print("❌ Error: 'openpyxl' is not installed. Run 'pip install openpyxl' first.")
        return

    # ✨ CHANGE 1: Determine the tag and ONLY delete old records for that specific tag
    pricing_tag = 'CASHLESS' if is_cashless else 'CASH'
    ServiceMaster.objects.filter(pricing_type=pricing_tag).delete()
    
    # Check if the file exists
    if not os.path.exists(filepath):
        print(f"⚠️ Notice: Could not find '{filepath}'. Skipping {pricing_tag} import.")
        return

    print(f"📂 Opening {pricing_tag} Excel file: {filepath}...")
    wb = openpyxl.load_workbook(filepath, data_only=True)
    
    count = 0

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"  ➡️ Reading sheet: {sheet_name}...")

        header_row_idx = None
        headers = []
        
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            if row and 'Description' in [str(c).strip() for c in row if c]:
                header_row_idx = row_idx
                headers = [str(c).strip() if c else '' for c in row]
                break

        if header_row_idx is None:
            print(f"     ⚠️ Warning: Could not find 'Description' header in sheet '{sheet_name}'. Skipping.")
            continue

        for row in sheet.iter_rows(min_row=header_row_idx + 2, values_only=True):
            if not row:
                continue

            row_dict = dict(zip(headers, row))

            desc = str(row_dict.get('Description', '')).strip()
            if not desc or desc == 'None':
                continue

            raw_rate = row_dict.get('Rate', 0)
            try:
                rate = float(raw_rate) if raw_rate else 0.0
            except (ValueError, TypeError):
                rate = 0.0

            code = str(row_dict.get('Code', '')).strip()
            if code == 'None': 
                code = ''

            cat = str(row_dict.get('Type of Service', '')).strip().upper()
            if not cat or cat == 'NONE': 
                cat = sheet_name.upper()

            # ✨ CHANGE 2: Save the pricing_tag to the database
            ServiceMaster.objects.create(
                category=cat,
                pricing_type=pricing_tag, # <-- Added this line
                description=desc,
                code=code,
                rate=rate
            )
            count += 1
            
    print(f"🎉 BOOM! {count} {pricing_tag} services successfully imported!\n")


if __name__ == '__main__':
    print("Starting Data Import Process...\n")
    
    # ✨ CHANGE 3: Run the import for BOTH files in your data folder
    
    # 1. Import Cashless Data (Using your existing file)
    run_import('./data/SANGIESIC.xlsx', is_cashless=True)
    
    # 2. Import Cash Data 
    # (When your senior provides the cash prices, save it in the data folder 
    # and update the filename below if it is named differently)
    run_import('./data/CASH_PRICES.xlsx', is_cashless=False)
    
    print("✅ All imports completed!")